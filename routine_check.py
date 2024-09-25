from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


def main():
    is_continue = input('输入q返回，按下任意键开始')
    if is_continue == 'q':
        return

    date = input('请输入检查日期（例如2023-11-30）：')

    # 加载常规检查（whole）.xlsx
    path_routine_check_whole = 'D:/routine_check_repository/export/' + date + '常规检查（whole）.xlsx'
    wb_routine_check_whole = load_workbook(path_routine_check_whole)
    ws_routine_check_whole = wb_routine_check_whole['Sheet1']

    # 加载常规检查基础表
    path_routine_check_base = 'D:/routine_check_repository/basic_tables/常规检查基础表.xlsx'
    wb_routine_check_base = load_workbook(path_routine_check_base)
    ws_routine_check_base = wb_routine_check_base['Sheet1']

    # 填充常规检查表
    t = 1
    for i in range(4, ws_routine_check_whole.max_row + 1):
        row = []
        if ws_routine_check_whole[f'H{i}'].value is not None or ws_routine_check_whole[f'L{i}'].value:
            row.append(t)
            row.append(ws_routine_check_whole[f'B{i}'].value)
            row.append(ws_routine_check_whole[f'C{i}'].value)
            row.append(ws_routine_check_whole[f'D{i}'].value)
            row.append(ws_routine_check_whole[f'E{i}'].value)
            row.append(ws_routine_check_whole[f'F{i}'].value)
            row.append(ws_routine_check_whole[f'G{i}'].value)
            row.append(ws_routine_check_whole[f'H{i}'].value)
            row.append(ws_routine_check_whole[f'I{i}'].value)
            row.append(ws_routine_check_whole[f'J{i}'].value)
            row.append(ws_routine_check_whole[f'K{i}'].value)
            row.append(ws_routine_check_whole[f'L{i}'].value)
            row.append(ws_routine_check_whole[f'M{i}'].value)
            row.append(ws_routine_check_whole[f'N{i}'].value)
            row.append(ws_routine_check_whole[f'O{i}'].value)

            ws_routine_check_base.append(row)
            t += 1

    # 居中对齐和边框
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 获取工作表中的最大列数
    max_column = ws_routine_check_base.max_column - 1

    # 从第四行开始，对所有单元格设置居中和全框线
    for row in ws_routine_check_base.iter_rows(min_row=4, max_col=max_column):
        for cell in row:
            cell.alignment = alignment_center
            cell.border = thin_border

    wb_routine_check_base.save('D:/routine_check_repository/export/' + date + '常规检查.xlsx')
    print('已保存至D:/routine_check_repository/export/' + date + '常规检查.xlsx')


if __name__ == '__main__':
    main()
