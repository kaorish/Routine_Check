from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
import re


def main():
    is_continue = input('输入q返回，按下任意键开始')
    if is_continue == 'q':
        return

    date = input('请输入检查日期（例如2023-11-30）：')
    is_test = input('是否为测试（1为是，0为否）：')

    # 加载常规检查（whole）基础表
    path_routine_check_whole_base = 'D:/routine_check_repository/basic_tables/常规检查（whole）基础表.xlsx'
    wb_routine_check_whole_base = load_workbook(path_routine_check_whole_base)
    ws_routine_check_whole_base = wb_routine_check_whole_base['Sheet1']

    # 加载无人情况表
    path_nobody = 'D:/routine_check_repository/basic_tables/无人情况.xlsx'
    wb_nobody = load_workbook(path_nobody)
    ws_nobody = wb_nobody['Sheet1']

    # 加载查寝情况表
    path_check_condition = 'D:/routine_check_repository/import/' + date + '查寝情况.xlsx'
    wb_check_condition = load_workbook(path_check_condition)
    ws_check_condition = wb_check_condition['Sheet1']

    # 加载大功率及安全隐患表
    path_high_power = 'D:/routine_check_repository/import/' + date + '大功率及安全隐患.xlsx'
    wb_high_power = load_workbook(path_high_power)
    ws_high_power = wb_high_power['Sheet1']

    # 给常规检查（whole）表写入检查日期
    date_changed = str(date).replace('-', '.')
    for i in range(4, ws_routine_check_whole_base.max_row + 1):
        ws_routine_check_whole_base[f'B{i}'].value = date_changed

    # 清洗数据，查寝情况表
    for i in range(2, ws_check_condition.max_row + 1):
        # 乱填无人
        if isinstance(ws_check_condition[f'C{i}'].value, str):
            ws_check_condition[f'C{i}'].value = '无人'

        # 楼栋不以大写F开头
        if str(ws_check_condition[f'A{i}'].value) not in ['F13', 'F14', 'F17']:
            number = re.findall('(\d+)', ws_check_condition[f'A{i}'].value)[0]
            ws_check_condition[f'A{i}'].value = 'F' + number

    # 填充常规检查（whole）表
    # 将查寝情况表的内容（卫生情况和理由）填入进去
    for i in range(2, ws_check_condition.max_row + 1):
        for j in range(4, ws_routine_check_whole_base.max_row + 1):
            # 楼栋和寝室号匹配成功，导入信息
            if ws_check_condition[f'A{i}'].value == ws_routine_check_whole_base[f'F{j}'].value and ws_check_condition[
                f'B{i}'].value == ws_routine_check_whole_base[f'G{j}'].value:
                # 若为无人则只填无人，否则填其他的所有信息
                if ws_check_condition[f'C{i}'].value == '无人':
                    ws_routine_check_whole_base[f'L{j}'].value = '无人'
                else:
                    # 分数
                    if ws_check_condition[f'C{i}'].value >= 80:
                        ws_routine_check_whole_base[f'H{j}'].value = '好'
                    elif ws_check_condition[f'C{i}'].value >= 60:
                        ws_routine_check_whole_base[f'H{j}'].value = '一般'
                    else:
                        ws_routine_check_whole_base[f'H{j}'].value = '差'
                    # 差寝理由1、2、3
                    ws_routine_check_whole_base[f'I{j}'].value = ws_check_condition[f'D{i}'].value
                    ws_routine_check_whole_base[f'J{j}'].value = ws_check_condition[f'E{i}'].value
                    ws_routine_check_whole_base[f'K{j}'].value = ws_check_condition[f'F{i}'].value
                # 直接匹配下一条，别浪费时间
                break

    # 把大功率及安全隐患表的内容加入进去
    for i in range(4, ws_high_power.max_row + 1):
        if ws_high_power[f'F{i}'].value is not None:
            for j in range(4, ws_routine_check_whole_base.max_row + 1):
                # 楼栋和寝室号匹配成功，导入信息
                if ws_high_power[f'D{i}'].value == ws_routine_check_whole_base[f'F{j}'].value and ws_high_power[
                    f'E{i}'].value == ws_routine_check_whole_base[f'G{j}'].value:
                    if '机' in ws_high_power[f'F{i}'].value or '烧' in ws_high_power[f'F{i}'].value or '热' in \
                            ws_high_power[f'F{i}'].value or '冰' in ws_high_power[f'F{i}'].value or '锅' in \
                            ws_high_power[f'F{i}'].value or '小' in ws_high_power[f'F{i}'].value or '电线' in  \
                            ws_high_power[f'F{i}'].value or '路由' in ws_high_power[f'F{i}'].value or '风扇' in \
                            ws_high_power[f'F{i}'].value or '卷发棒' in ws_high_power[f'F{i}'].value \
                            or '器' in ws_high_power[f'F{i}'].value:
                        ws_routine_check_whole_base[f'N{j}'].value = ws_high_power[f'F{i}'].value
                    # 直接匹配下一条，别浪费时间
                    break

    # 把无人情况记录在无人情况表中
    def update_nobody():
        # 记录无人情况
        for i in range(2, ws_check_condition.max_row + 1):
            if ws_check_condition[f'C{i}'].value == '无人':
                for j in range(2, ws_nobody.max_row + 1):
                    if ws_check_condition[f'A{i}'].value == ws_nobody[f'A{j}'].value and ws_check_condition[
                        f'B{i}'].value == ws_nobody[f'B{j}'].value:
                        ws_nobody[f'C{j}'].value = int(ws_nobody[f'C{j}'].value) + 1
                        break

    # 是否为测试，毕竟别把无人次数多记了
    if is_test == 0:
        update_nobody()
        print('已将无人情况添加至 D:/routine_check_repository/basic_tables/无人情况.xlsx')

    # 居中对齐和边框
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 获取工作表中的最大列数
    max_column = ws_routine_check_whole_base.max_column - 1

    # 从第四行开始，对所有单元格设置居中和全框线
    for row in ws_routine_check_whole_base.iter_rows(min_row=4, max_col=max_column):
        for cell in row:
            cell.alignment = alignment_center
            cell.border = thin_border

    wb_routine_check_whole_base.save('D:/routine_check_repository/export/' + date + "常规检查（whole）.xlsx")
    print('已保存至D:/routine_check_repository/export/' + date + "常规检查（whole）.xlsx")

if __name__ == '__main__':
    main()
