from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side


def main():
    is_continue = input('输入q返回，按下任意键开始')
    if is_continue == 'q':
        return

    date = input('请输入检查日期（例如2023-11-30）：')
    week = input('这是第几周：')

    # 加载宿舍通报基础表
    path_notification_base = 'D:/routine_check_repository/basic_tables/宿舍通报基础表.xlsx'
    wb_notification_base = load_workbook(path_notification_base)
    ws_notification_base_good = wb_notification_base['优秀寝室']
    ws_notification_base_bad = wb_notification_base['不文明寝室']

    # 加载无人情况表
    path_nobody = 'D:/routine_check_repository/basic_tables/无人情况.xlsx'
    wb_nobody = load_workbook(path_nobody)
    ws_nobody = wb_nobody['Sheet1']

    # 加载住宿信息
    path_information = 'D:/routine_check_repository/basic_tables/住宿信息.xlsx'
    wb_information = load_workbook(path_information)
    ws_information = wb_information['Sheet1']

    # 加载优秀寝室
    path_good = 'D:/routine_check_repository/import/第' + week + '周优秀寝室.xlsx'
    wb_good = load_workbook(path_good)
    ws_good = wb_good['Sheet1']

    # 加载常规检查表
    path_routine_check = 'D:/routine_check_repository/export/' + date + '常规检查.xlsx'
    wb_routine_check = load_workbook(path_routine_check)
    ws_routine_check = wb_routine_check['Sheet1']

    # 添加优秀寝室
    t = 1
    for i in range(2, ws_good.max_row + 1):
        row = []
        for j in range(2, ws_information.max_row + 1):
            if ws_good[f'A{i}'].value == ws_information[f'B{j}'].value and ws_good[f'B{i}'].value == ws_information[f'C{j}'].value:
                row.append(t)
                row.append(ws_information[f'B{j}'].value)
                row.append(ws_information[f'C{j}'].value)
                row.append(ws_information[f'D{j}'].value)
                row.append(ws_information[f'E{j}'].value)
                row.append(ws_information[f'F{j}'].value)
                row.append(ws_information[f'G{j}'].value)

                ws_notification_base_good.append(row)
                t += 1
                break

    # 添加不文明寝室--卫生为差
    t = 1
    for i in range(4, ws_routine_check.max_row + 1):
        row = []
        if ws_routine_check[f'H{i}'].value == '差':
            for j in range(2, ws_information.max_row + 1):
                if ws_routine_check[f'F{i}'].value == ws_information[f'B{j}'].value and ws_routine_check[f'G{i}'].value == ws_information[
                    f'C{j}'].value:
                    row.append(t)
                    row.append(ws_information[f'B{j}'].value)
                    row.append(ws_information[f'C{j}'].value)
                    row.append(ws_information[f'D{j}'].value)
                    row.append(ws_information[f'E{j}'].value)
                    row.append(ws_information[f'F{j}'].value)
                    row.append(ws_information[f'G{j}'].value)

                    ws_notification_base_bad.append(row)
                    t += 1
                    break

    # 添加不文明寝室--满三次无人情况
    for i in range(2, ws_nobody.max_row + 1):
        row = []
        if ws_nobody[f'C{i}'].value == 3:
            ws_nobody[f'C{i}'].value = 0
            for j in range(2, ws_information.max_row + 1):
                if ws_nobody[f'A{i}'].value == ws_information[f'B{j}'].value and ws_nobody[f'B{i}'].value == ws_information[
                    f'C{j}'].value:
                    row.append(t)
                    row.append(ws_information[f'B{j}'].value)
                    row.append(ws_information[f'C{j}'].value)
                    row.append(ws_information[f'D{j}'].value)
                    row.append(ws_information[f'E{j}'].value)
                    row.append(ws_information[f'F{j}'].value)
                    row.append(ws_information[f'G{j}'].value)

                    ws_notification_base_bad.append(row)
                    t += 1
                    break

    # 居中对齐和边框
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # 获取工作表中的最大列数
    max_column_good = ws_notification_base_good.max_column
    max_column_bad = ws_notification_base_bad.max_column

    # 从第二行开始，对所有单元格设置居中和全框线
    for row in ws_notification_base_good.iter_rows(min_row=2, max_col=max_column_good):
        for cell in row:
            cell.alignment = alignment_center
            cell.border = thin_border

    for row in ws_notification_base_bad.iter_rows(min_row=2, max_col=max_column_bad):
        for cell in row:
            cell.alignment = alignment_center
            cell.border = thin_border

    wb_notification_base.save('D:/routine_check_repository/export/第' + week + '周宿舍通报.xlsx')
    print('已保存至D:/routine_check_repository/export/第' + week + '周宿舍通报.xlsx')

if __name__ == '__main__':
    main()
